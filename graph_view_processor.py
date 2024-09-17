from tkinter import filedialog, messagebox
import numpy as np
import matplotlib.widgets
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button


def removeLast(string, end):
    string = string[:len(string) - end]
    return string

class graph_processor:
    def file_loader(self, gui):
        try:
            self.file_name = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel file", "*.xlsx")])
            self.wb = load_workbook(self.file_name)
            self.ws = self.wb.active
            gui.initial_slider_value_1 = 0
            gui.initial_slider_value_2 = 0
            gui.viewer_open_x = []
            gui.viewer_open_y = []
            gui.viewer_close_x = []
            gui.viewer_close_y = []
            gui.viewer_normalized_open_x = []
            gui.viewer_normalized_open_y = []
            gui.viewer_normalized_close_x = []
            gui.viewer_normalized_close_y = []
            gui.lines1.set_xdata(gui.viewer_open_x)
            gui.lines1.set_ydata(gui.viewer_open_y)
            gui.lines2.set_xdata(gui.viewer_normalized_open_x)
            gui.lines2.set_ydata(gui.viewer_normalized_open_y)
            gui.lines3.set_xdata(gui.viewer_close_x)
            gui.lines3.set_ydata(gui.viewer_close_y)
            gui.lines4.set_xdata(gui.viewer_normalized_close_x)
            gui.lines4.set_ydata(gui.viewer_normalized_close_y)
            gui.canvas1.draw()
            self.column_names = ["A17", "C17", "E17", "G17"]
            self.current_column = ["A19", "C19", "E19", "G19"]
            self.next_column = ["B19", "D19", "F19", "H19"]
            for item in range(4):
                if self.ws[self.column_names[item]].value == "Open Aperture":
                    range1 = self.ws[f"{self.current_column[item]}:{self.current_column[item][0]}{self.ws.max_row}"]
                    range2 = self.ws[f"{self.next_column[item]}:{self.next_column[item][0]}{self.ws.max_row}"]
                    for cell in range1:
                        for x in cell:
                            gui.viewer_open_x.append(x.value)
                    for cell in range2:
                        for y in cell:
                            gui.viewer_open_y.append(y.value)
                elif self.ws[self.column_names[item]].value == "Close Aperture":
                    range1 = self.ws[
                        f"{self.current_column[item]}:{self.current_column[item][0]}{self.ws.max_row}"]
                    range2 = self.ws[f"{self.next_column[item]}:{self.next_column[item][0]}{self.ws.max_row}"]
                    for cell in range1:
                        for x in cell:
                            gui.viewer_close_x.append(x.value)
                    for cell in range2:
                        for y in cell:
                            gui.viewer_close_y.append(y.value)
                elif self.ws[self.column_names[item]].value == "Normalized Open Aperture":
                    range1 = self.ws[
                        f"{self.current_column[item]}:{self.current_column[item][0]}{self.ws.max_row}"]
                    range2 = self.ws[f"{self.next_column[item]}:{self.next_column[item][0]}{self.ws.max_row}"]
                    for cell in range1:
                        for x in cell:
                            gui.viewer_normalized_open_x.append(x.value)
                    for cell in range2:
                        for y in cell:
                            gui.viewer_normalized_open_y.append(y.value)
                elif self.ws[self.column_names[item]].value == "Normalized Close Aperture":
                    range1 = self.ws[
                        f"{self.current_column[item]}:{self.current_column[item][0]}{self.ws.max_row}"]
                    range2 = self.ws[f"{self.next_column[item]}:{self.next_column[item][0]}{self.ws.max_row}"]
                    for cell in range1:
                        for x in cell:
                            gui.viewer_normalized_close_x.append(x.value)
                    for cell in range2:
                        for y in cell:
                            gui.viewer_normalized_close_y.append(y.value)
            if len(gui.viewer_open_x) > 0:
                gui.lines1.set_xdata(gui.viewer_open_x)
                gui.lines1.set_ydata(gui.viewer_open_y)
                gui.ax1.set_ylim([min(gui.viewer_open_y) - (min(gui.viewer_open_y) * 0.01),
                                  max(gui.viewer_open_y) + (max(gui.viewer_open_y) * 0.01)])
                gui.ax1.set_xlim([min(gui.viewer_open_x), max(gui.viewer_open_x)])
                gui.normalize_button_1["state"] = "normal"
            if len(gui.viewer_close_x) > 0:
                gui.lines3.set_xdata(gui.viewer_close_x)
                gui.lines3.set_ydata(gui.viewer_close_y)
                gui.ax3.set_ylim([min(gui.viewer_close_y) - (min(gui.viewer_close_y) * 0.05),
                                  max(gui.viewer_close_y) + (max(gui.viewer_close_y) * 0.05)])
                gui.ax3.set_xlim([min(gui.viewer_close_x), max(gui.viewer_close_x)])
                gui.normalize_button_2["state"] = "normal"
            if len(gui.viewer_normalized_open_x) > 0:
                gui.lines2.set_xdata(gui.viewer_normalized_open_x)
                gui.lines2.set_ydata(gui.viewer_normalized_open_y)
                gui.ax2.set_ylim([min(gui.viewer_normalized_open_y) - (min(gui.viewer_normalized_open_y) * 0.01),
                                  max(gui.viewer_normalized_open_y) + (max(gui.viewer_normalized_open_y) * 0.01)])
                gui.ax2.set_xlim([min(gui.viewer_normalized_open_x), max(gui.viewer_normalized_open_x)])
            if len(gui.viewer_normalized_close_x) > 0:
                gui.lines4.set_xdata(gui.viewer_normalized_close_x)
                gui.lines4.set_ydata(gui.viewer_normalized_close_y)
                gui.ax4.set_ylim([min(gui.viewer_normalized_close_y) - (min(gui.viewer_normalized_close_y) * 0.01),
                                  max(gui.viewer_normalized_close_y) + (max(gui.viewer_normalized_close_y) * 0.01)])
                gui.ax4.set_xlim([min(gui.viewer_normalized_close_x), max(gui.viewer_normalized_close_x)])
            gui.canvas1.draw()
            try:
                gui.calculation.wavelength_value.config(
                    text=f"{removeLast(self.ws["C5"].value, 3)} {gui.calculation.wavelength_unit}")
            except:
                pass
            try:
                gui.calculation.laser_beam_input_power_value.config(
                    text=f"{removeLast(self.ws["C6"].value, 2)} {gui.calculation.laser_beam_input_power_unit}")
            except:
                pass
            try:
                gui.calculation.laser_beam_diameter_value.config(
                    text=f"{removeLast(self.ws["C7"].value, 3)} {gui.calculation.laser_beam_diameter_unit}")
            except:
                pass
            try:
                gui.calculation.focal_length_value.config(
                    text=f"{removeLast(self.ws["C8"].value, 3)} {gui.calculation.self.focal_length_unit}")
            except:
                pass
            try:
                gui.calculation.close_aperture_value.config(
                    text=f"{removeLast(self.ws["C9"].value, 3)} {gui.calculation.close_aperture_unit}")
            except:
                pass
            try:
                gui.calculation.beam_rad_at_aperture_value.config(
                    text=f"{removeLast(self.ws["C10"].value, 3)} {gui.calculation.beam_rad_at_aperture_unit}")
            except:
                pass
            try:
                gui.calculation.z_scan_sample_thickness_value.config(
                    text=f"{removeLast(self.ws["C11"].value, 3)} {gui.calculation.z_scan_sample_thickness_unit}")
            except:
                pass
            try:
                gui.calculation.linear_refractive_index_value.config(
                    text=self.ws["C12"].value)
            except:
                pass
            try:
                gui.calculation.transmittance_value.config(
                    text=f"{removeLast(self.ws["C13"].value, 2)} {gui.calculation.transmittance_unit}")
            except:
                pass
            try:
                gui.calculation.transmittance_thickness_value.config(
                    text=f"{removeLast(self.ws["C14"].value, 3)} {gui.calculation.transmittance_thickness_unit}")
            except:
                pass
            try:
                gui.title_value = self.ws["A1"].value
                gui.sweeping_distance_value = self.ws["C2"].value
                gui.step_distance_value = self.ws["C3"].value
                gui.bandwidth_value = self.ws["C4"].value
            except:
                pass
            gui.calculation.update_button["state"] = "disabled"
        except:
            messagebox.showerror("File error", "Please upload a valid format")


    def normalize(self, gui, number):
        self.edge_fraction_1 = 0
        self.edge_fraction_2 = 0
        x=[]
        y=[]
        self.slider: matplotlib.widgets.Slider
        if number == 1:
            x = gui.viewer_open_x
            y = gui.viewer_open_y
        elif number == 2:
            x = gui.viewer_close_x
            y = gui.viewer_close_y

        fig, ax = plt.subplots()
        plt.subplots_adjust(left=0.1, bottom=0.30)
        plt.plot(x, y)
        min_val = min(y)
        max_val = max(y)
        ln = plt.axhline(y=(min_val + max_val) / 2, color='r', linestyle='--')
        plt.xlabel('Distance in mm')
        plt.ylabel('Power in watt')

        axcolor = 'lightgoldenrodyellow'
        slider_config = plt.axes((0.1, 0.15, 0.8, 0.03), facecolor=axcolor)
        button_config = plt.axes((0.45, 0.01, 0.1, 0.07), facecolor=axcolor)

        if number == 1 and len(gui.viewer_normalized_open_x) == 0 and gui.initial_slider_value_1 == 0:
            gui.initial_slider_value_1 = (min_val + max_val) / 2
            self.slider = Slider(slider_config, 'Normalize', min_val, max_val, valinit=gui.initial_slider_value_1)
        elif number == 2 and len(gui.viewer_normalized_close_x) == 0 and gui.initial_slider_value_2 == 0:
            gui.initial_slider_value_2 = (min_val + max_val) / 2
            self.slider = Slider(slider_config, 'Normalize', min_val, max_val, valinit=gui.initial_slider_value_2)
        elif number == 1:
            self.slider = Slider(slider_config, 'Normalize', min_val, max_val, valinit=gui.initial_slider_value_1)
        elif number == 2:
            self.slider = Slider(slider_config, 'Normalize', min_val, max_val, valinit=gui.initial_slider_value_2)

        button = Button(button_config, "Update")

        def update(val):
            if number == 1:
                self.edge_fraction_1 = self.slider.val
                ln.set_ydata(self.edge_fraction_1)
                gui.initial_slider_value_1 = self.edge_fraction_1
            elif number == 2:
                self.edge_fraction_2 = self.slider.val
                ln.set_ydata(self.edge_fraction_2)
                gui.initial_slider_value_2 = self.edge_fraction_2
            fig.canvas.draw_idle()

        def apply(val):
            if number ==1:
                gui.viewer_normalized_open_x = gui.viewer_open_x
                gui.viewer_open_y = np.array(gui.viewer_open_y)
                gui.viewer_normalized_open_y = gui.viewer_open_y / gui.initial_slider_value_1
                gui.lines2.set_xdata(gui.viewer_normalized_open_x)
                gui.lines2.set_ydata(gui.viewer_normalized_open_y)
                gui.ax2.set_ylim([min(gui.viewer_normalized_open_y) - (min(gui.viewer_normalized_open_y) * 0.01),
                                  max(gui.viewer_normalized_open_y) + (max(gui.viewer_normalized_open_y) * 0.01)])
                gui.ax2.set_xlim([min(gui.viewer_normalized_open_x), max(gui.viewer_normalized_open_x)])
                gui.calculation.del_phi_close_value.config(text=1-min(gui.viewer_normalized_open_y))
            if number == 2:
                gui.viewer_normalized_close_x = gui.viewer_close_x
                gui.viewer_close_y = np.array(gui.viewer_close_y)
                gui.viewer_normalized_close_y = gui.viewer_close_y / gui.initial_slider_value_2
                gui.lines4.set_xdata(gui.viewer_normalized_close_x)
                gui.lines4.set_ydata(gui.viewer_normalized_close_y)
                gui.ax4.set_ylim([min(gui.viewer_normalized_close_y) - (min(gui.viewer_normalized_close_y) * 0.05),
                                  max(gui.viewer_normalized_close_y) + (max(gui.viewer_normalized_close_y) * 0.05)])
                gui.ax4.set_xlim([min(gui.viewer_normalized_close_x), max(gui.viewer_normalized_close_x)])
                gui.calculation.del_phi_open_value.config(text=max(gui.viewer_normalized_close_y) - min(gui.viewer_normalized_close_y))
            gui.canvas1.draw()

        self.slider.on_changed(update)
        button.on_clicked(apply)

        plt.show()

